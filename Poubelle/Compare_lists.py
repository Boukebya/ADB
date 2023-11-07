import pandas as pd 
import openpyxl

#obtenir les valeurs des résultats
file_res = r"Data_fourniture_test_rés2.xlsx"

data_res = pd.ExcelFile(file_res)

name_sheet_res = data_res.sheet_names[0]

ps_res = openpyxl.load_workbook(file_res)

sheet_res = ps_res[name_sheet_res]

quantité_res = []
référence_res = []

for row in range(2, sheet_res.max_row + 1):

    # each row in the spreadsheet represents information for a particular purchase.
    quantité_res.append(sheet_res["A" + str(row)].value)

    référence_res.append(sheet_res["B" + str(row)].value)

# print(quantité_res)
# print(référence_res)



#obtenir les valeurs des résultats
file_devis = r"Data_fourniture_test_devis2.xlsx"

data_devis = pd.ExcelFile(file_devis)

name_sheet_devis = data_devis.sheet_names[0]

ps_devis = openpyxl.load_workbook(file_devis)

sheet_devis = ps_devis[name_sheet_devis]

quantité_devis = []
référence_devis = []

for row in range(2, sheet_devis.max_row + 1):

    # each row in the spreadsheet represents information for a particular purchase.
    quantité_devis.append(sheet_devis["A" + str(row)].value)

    référence_devis.append(sheet_devis["B" + str(row)].value)

# print(quantité_devis)
# print(référence_devis)

conf_q = 0 
conf_ref = 0

i = 0
j = 0
for elem in référence_devis:

    if elem in référence_res:
        print("référence : %s" % elem)
        conf_ref+=1

        index = référence_res.index(elem)
        j+=1

        if quantité_devis[i] == quantité_res[index]:
            conf_q+=1
            

    i+=1


#La confiance sur les références dépend de tous les objets perçus dans la liste de devis
conf_ref= (conf_ref/i)*100

#Alors que la confiance sur les quantités ne se calcule que si l'on trouve déjà la référence produit
conf_q = (conf_q/j)*100

print("confiance quantité: %s" % conf_q,"%")
print("confiance référence: %s" % conf_ref,"%")
